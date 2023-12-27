from tkinter import *
from tkinter import ttk
from DES import encryptDES,decryptDES
from math import ceil
from random import randint
from tkinter.filedialog import askopenfile,asksaveasfile
import docx
from openpyxl import load_workbook
import io

#config
FontBig = ('Arial',15,'bold')
FontSmall = ('Arial',13,'bold')
BG = '#CCCCFF'

window = Tk()
window.title(' Ứng dụng thuật toán DES bảo mật đề thi tuyển sinh.')
window.geometry('1600x700')
window.config(bg=BG)

notebook = ttk.Notebook(window)
notebook.pack(pady=0)

frame1 = Frame(notebook, width=1600, height=700, padx=20, pady=20)
frame2 = Frame(notebook, width=1600, height=700,  padx=20, pady=20)

frame1.pack(fill='both', expand=1)
frame2.pack(fill='both', expand=1)

notebook.add(frame1, text='Mã hóa')
notebook.add(frame2, text='Giải mã')

key_label = Label(frame1, text='Khóa bí mật:', font=FontBig)
key_label.place(x=80, y=40)

source1_label = Label(frame1, text='Nhập bản rõ:', font=FontBig)
source1_label.place(x=0, y=130)

ans1_label = Label(frame1, text='Bản mã:', font=FontBig)
ans1_label.place(x=870, y=130)

key_decrypt_label = Label(frame2, text='Khóa bí mật: ', font=FontBig)
key_decrypt_label.place(x=80, y=40)

source2_label = Label(frame2, text='Nhập bản mã:', font=FontBig)
source2_label.place(x=0, y=130)

ans2_label = Label(frame2, text='Bản rõ:', font=FontBig)
ans2_label.place(x=870, y=130)

key_entry = Entry(frame1, width=20, font=FontBig)
key_entry.place(x=240, y=40)

source1_entry = Text(frame1, width=85, height=17)
source1_entry.place(x=0, y=170)

ans1_entry = Text(frame1, width=85, height=17)
ans1_entry.place(x=870, y=170)

key_decrypt_entry = Entry(frame2, width=20, font=FontBig)
key_decrypt_entry.place(x=240, y=40)

source2_entry = Text(frame2, width=85, height=17)
source2_entry.place(x=0, y=170)

ans2_entry = Text(frame2, width=85, height=17)
ans2_entry.place(x=870, y=170)

def randomkey():
    key = randint(10000000,99999999)
    key_entry.delete(0,'end')
    key_entry.insert(0,str(key))


def readWordfile(filename):
    doc = docx.Document(filename)
    result = [p.text for p in doc.paragraphs]
    return '\n'.join(result)


def readExcelfile(filename):
    book = load_workbook(filename)
    sheet = book.active
    ans = ''
    for i in range(1,sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            ans+= str(sheet.cell(row=i, column=j).value)
            ans+=' '
        ans+='\n'
    return ans

def readTextfile(filename):
    s= ''
    file = io.open(filename, mode='r', encoding='utf-8')
    read = file.readlines()
    for line in read:
        s+=line
    return s

filetypes = [('Word Files', '*.docx'), ('Text Files', '*.txt'), ('Excel Files', '*.xlsx')]

def open():
    global data
    file = askopenfile(mode='r', initialdir='C:/Users/Admin/Desktop/An Toan Va Bao Mat Thong Tin/BTL', title='Select File', filetypes=filetypes)
    s = str(file)
    a = s.find('name')
    b = s.find('mode')
    filename = s[a+6:b-2]
    if 'docx' in filename:
        data = readWordfile(filename)
    elif 'xlsx' in filename:
        data = readExcelfile(filename)
    elif 'txt' in filename:
        data = readTextfile(filename)
    source1_entry.delete(1.0,END)
    source1_entry.insert(INSERT,data)

def decrypt_open():
    global data
    file = askopenfile(mode='r', initialdir='C:/Users/Admin/Desktop/An Toan Va Bao Mat Thong Tin/BTL',
                       title='Select File', filetypes=filetypes)
    s = str(file)
    a = s.find('name')
    b = s.find('mode')
    filename = s[a + 6:b - 2]
    if 'docx' in filename:
        data = readWordfile(filename)
    elif 'xlsx' in filename:
        data = readExcelfile(filename)
    elif 'txt' in filename:
        data = readTextfile(filename)
    source2_entry.delete(1.0, END)
    source2_entry.insert(INSERT, data)


def save():
    f = asksaveasfile(mode='w', defaultextension=".txt")
    if f is None:
        return
    text2save = str(ans1_entry.get(1.0, END))
    f.write(text2save)
    f.close()


def decrypt_save():
    f = asksaveasfile(mode='wb', defaultextension=".txt")
    if f is None:
        return
    text2save = str(ans2_entry.get(1.0, END)).encode('utf8')
    f.write(text2save)
    f.close()

def encrypt():
    key = key_entry.get()
    text = source1_entry.get(1.0, END)
    if len(key) == 8:
        ans = encryptDES(key, text)
        ans1_entry.delete(1.0, END)
        ans1_entry.insert(INSERT, ans)
    else:
        var = StringVar()
        label_message = Message(frame1, width=200, padx=200, pady=100, textvariable=var, relief=RAISED, bg='#FD0011', fg='#FFFFFF',font=20)
        label_message.pack()
        var.set("Khóa bí mật cần có độ dài 8 kí tự")
        label_message.after(800, label_message.destroy)


def decrypt():
    key = key_decrypt_entry.get()
    text = source2_entry.get(1.0, END)
    ans = decryptDES(key, text)
    ans2_entry.delete(1.0, END)
    ans2_entry.insert(INSERT, ans)


def exit():
    window.destroy()


rd_key_btn = Button(frame1, text='Tạo ngẫu nhiên', width=15, height=1, font=('Arial', 10, 'bold'), command=randomkey)
rd_key_btn.place(x=480, y=40)

upfile1_btn = Button(frame1, text='Đọc từ file', width=15, height=1, font=('Arial', 10, 'bold'), command=open)
upfile1_btn.place(x=250, y=450)

savefile1_btn = Button(frame1, text='Lưu vào file', width=15, height=1, font=('Arial', 10, 'bold'), command=save)
savefile1_btn.place(x=1180, y=450)

encrypt_btn = Button(frame1, text='Mã hóa', width=30, height=2, font=('Arial', 15, 'bold'), command=encrypt)
encrypt_btn.place(x=350, y=550)

exit1_btn = Button(frame1, text='Thoát', width=30, height=2, font=('Arial', 15, 'bold'), command=exit)
exit1_btn.place(x=800, y=550)

upfile2_btn = Button(frame2, text='Đọc từ file', width=15, height=1, font=('Arial', 10, 'bold'), command=decrypt_open)
upfile2_btn.place(x=250, y=450)

savefile2_btn = Button(frame2, text='Lưu vào file', width=15, height=1, font=('Arial', 10, 'bold'), command=decrypt_save)
savefile2_btn.place(x=1180, y=450)

decrypt_btn = Button(frame2, text='Giải mã', width=30, height=2, font=('Arial', 15, 'bold'), command=decrypt)
decrypt_btn.place(x=350, y=550)

exit1_btn = Button(frame2, text='Thoát', width=30, height=2, font=('Arial', 15, 'bold'), command=exit)
exit1_btn.place(x=800, y=550)

window.mainloop()
